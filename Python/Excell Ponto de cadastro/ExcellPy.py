from openpyxl import Workbook, load_workbook
import time
import os
print("============Olá, Sistema de Ponto 0.01============")
user = os.getlogin()
login = input("\n Login:")
senha = input(" Senha:")
os.system('cls') or None
osponto = login + '.xlsx'
if (os.path.isfile(osponto)):
    ponto = load_workbook(osponto)
else:
    ponto = Workbook()
planilha1 = ponto.active
dia = time.strftime('%d/%m/%y')
hr = time.strftime('%H:%M')
esc = int(input("Digite o numero correspondente\n1-Entrada \n2-Saída\n"))
if esc == 1:
    sep = ["————————Entrada————————"]
elif esc == 2:
    sep = ["—————————Saída—————————"]
else:
    ZZZ = input("OPÇÃO INVALIDA, REINICIE A APLICAÇÃO")
nome = ["Nome:", login]
hora = ["Hora", hr]
hj = ["Dia:", dia]
planilha1.append(sep)
planilha1.append(nome)
planilha1.append(hora)
planilha1.append(hj)
ponto.save(login + ".xlsx")